!pip install transformers[torch] accelerate -U
!pip install openpyxl scikit-learn umap-learn

import os
import pandas as pd
import torch

from transformers import BertTokenizer, BertModel
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
import umap
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import logging

# Configuration - the file paths will need to be changed when someone else run's the code
FILE_PATH = '/content/Original Hourly Billing Activity Details.xlsx'
OUTPUT_DIR = '/content/'

# Logging setup
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Check for GPU availability
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
logging.info(f"Using device: {device}")

# Function to load data
def load_data(file_path):
    try:
        data = pd.read_excel(file_path, sheet_name='Export')
        return data
    except Exception as e:
        logging.error(f"Error loading data: {e}")
        raise

# Function to preprocess data
def preprocess_data(data):
    data['Hours(Duration)'] = data.apply(
        lambda row: row['Billable Amount'] / row['Hourly Rate'] if pd.isnull(row['Hours(Duration)']) else row['Hours(Duration)'],
        axis=1
    )
    data['Text'] = data['Service Type'].astype(str) + " " + data['Description'].astype(str)
    data['Text'] = data['Text'].fillna('').astype(str)
    data = data[~data['Text'].str.contains('nan nan')]
    return data

# Function to generate BERT embeddings
def generate_embeddings(texts, tokenizer, model, device, batch_size=32):
    embeddings = []
    for i in range(0, len(texts), batch_size):
        batch_texts = texts[i:i+batch_size].tolist()
        inputs = tokenizer(batch_texts, return_tensors="pt", max_length=512, truncation=True, padding=True)
        inputs = {k: v.to(device) for k, v in inputs.items()}
        with torch.no_grad():
            outputs = model(**inputs)
        batch_embeddings = outputs.last_hidden_state[:, 0, :].cpu().numpy()
        embeddings.append(batch_embeddings)
    embeddings = np.vstack(embeddings)
    return embeddings

# Function to perform clustering
def perform_clustering(features, num_clusters=50):
    scaler = StandardScaler()
    features_scaled = scaler.fit_transform(features)
    umap_model = umap.UMAP(n_components=10, random_state=42)
    features_reduced = umap_model.fit_transform(features_scaled)
    kmeans = KMeans(n_clusters=num_clusters, random_state=42)
    clusters = kmeans.fit_predict(features_reduced)
    return clusters

# Function to map categories and rank automation ease
def map_categories_and_rank(data):
    category_names = {
        0: "Bank Feeds Processing",
        1: "Expense Reimbursement",
        2: "Accounts Payable",
        3: "Client Communication",
        4: "Invoice Processing",
        5: "Financial Reporting",
        6: "Payroll Processing and Reconciliation",
        7: "Vendor Management",
        8: "Month-End Close",
        9: "General Administration",
        10: "Audit Support and Compliance",
        11: "AR Management",
        12: "Employee Benefits Administration",
        13: "Budgeting and Forecasting",
        14: "Cash Management and Flow",
        15: "Tax Preparation and Filings",
        16: "Data Entry",
        17: "Compliance",
        18: "Training and Onboarding",
        19: "Strategic Advisory",
        20: "Financial Analysis",
        21: "Financial Modeling and Analysis",
        22: "Operational Management",
        23: "Consulting",
        24: "Tax Advisory",
        25: "Financial Planning",
        26: "Regulatory Reporting",
        27: "Investment Management",
        28: "Project Management",
        29: "Risk Management and Assessment",
        30: "Client Day-to-Day Touchpoint",
        31: "Client Communication and Correspondence",
        32: "Vendor Bill Processing and Communication",
        33: "AR Deposits and Review",
        34: "Expense Entries and Reconciliation",
        35: "General Ledger Review",
        36: "Financial Statements Preparation",
        37: "Audit and Compliance Support",
        38: "Training and Support for New Employees",
        39: "Monthly and Quarterly Close",
        40: "Revenue Recognition",
        41: "Budget Preparation and Review",
        42: "Cash Flow Management",
        43: "Employee Benefits Administration",
        44: "Financial Modeling and Analysis",
        45: "Investment Advisory",
        46: "Risk Assessment and Management",
        47: "Strategic Financial Planning"
    }

    data['Category'] = data['Cluster'].map(category_names)

    # Define automation ease
    automation_ease = {
        "Bank Feeds Processing": 5,
        "Accounts Payable": 5,
        "Invoice Processing": 5,
        "Cash Management": 5,
        "Tax Preparation": 5,
        "Data Entry": 5,
        "Expense Reimbursement": 4,
        "Payroll Processing": 4,
        "Vendor Management": 4,
        "General Administration": 4,
        "AR Management": 4,
        "Employee Benefits": 4,
        "Compliance": 4,
        "Financial Modeling": 4,
        "Operational Management": 4,
        "Tax Advisory": 4,
        "Financial Planning": 4,
        "Regulatory Reporting": 4,
        "Investment Management": 4,
        "Project Management": 4,
        "Risk Management": 4,
        "Client Communication": 2,
        "Financial Reporting": 2,
        "Month-End Close": 2,
        "Audit Support": 2,
        "Budgeting and Forecasting": 2,
        "Training and Onboarding": 2,
        "Strategic Advisory": 2,
        "Financial Analysis": 2,
        "Consulting": 2
    }

    data['Automation Ease'] = data['Category'].map(automation_ease).fillna(3)  # Default to 3 for categories not explicitly listed
    return data

# Main execution flow
def main():
    # Load data
    data = load_data(FILE_PATH)

    # Preprocess data
    data = preprocess_data(data)
    logging.info(f"Total hours after preprocessing: {data['Hours(Duration)'].sum()}")

    # Load BERT tokenizer and model
    tokenizer = BertTokenizer.from_pretrained('bert-base-uncased')
    model = BertModel.from_pretrained('bert-base-uncased')
    model.to(device)

    # Generate embeddings
    embeddings = generate_embeddings(data['Text'], tokenizer, model, device)

    # Generate TF-IDF vectors
    vectorizer = TfidfVectorizer(max_features=5000)
    tfidf_matrix = vectorizer.fit_transform(data['Text'])

    # Combine BERT embeddings and TF-IDF vectors
    combined_features = np.hstack((embeddings, tfidf_matrix.toarray()))

    # Perform clustering
    data['Cluster'] = perform_clustering(combined_features)

    # Save the clusters to inspect the categories
    clustered_data = data[['Client Name', 'Operator', 'Cluster', 'Text', 'Hours(Duration)']]
    clustered_data.to_excel(os.path.join(OUTPUT_DIR, 'Clustered_Data.xlsx'), index=False)
    logging.info("Clustered data saved for inspection.")

    # Check total hours after clustering
    logging.info(f"Total hours after clustering: {data['Hours(Duration)'].sum()}")

    # Map categories and rank automation ease
    data = map_categories_and_rank(data)

    # Cost Analysis
    data['Cost Per Hour'] = data['Billable Amount'] / data['Hours(Duration)']
    category_costs = data.groupby('Category')['Billable Amount'].sum().reset_index()
    category_costs.rename(columns={'Billable Amount': 'Total Cost'}, inplace=True)
    data = pd.merge(data, category_costs, on='Category', how='left')

    # Save the initial data for inspection
    initial_data_file = os.path.join(OUTPUT_DIR, 'Initial_Data_With_Categories_And_Automation_Ease.xlsx')
    data.to_excel(initial_data_file, index=False)
    logging.info(f"Initial data with categories and automation ease saved as {initial_data_file}")

    # Group and summarize data
    grouped = data.groupby(['Client Name', 'Operator', 'Category', 'Automation Ease']).agg({'Hours(Duration)': 'sum', 'Billable Amount': 'sum', 'Total Cost': 'first'}).reset_index()
    grouped['Cost Per Hour'] = grouped['Billable Amount'] / grouped['Hours(Duration)']
    grouped = grouped.sort_values(by=['Client Name', 'Hours(Duration)', 'Automation Ease'], ascending=[True, False, True])

    summary = grouped.groupby(['Category', 'Automation Ease']).agg({'Hours(Duration)': 'sum', 'Billable Amount': 'sum', 'Total Cost': 'first'}).reset_index()

    # Debug: Check total hours in summary
    logging.info(f"Total hours in summary: {summary['Hours(Duration)'].sum()}")

    # Save the summary data for inspection
    summary_file = os.path.join(OUTPUT_DIR, 'Summary_Data_Updated.xlsx')
    summary.to_excel(summary_file, index=False)
    logging.info(f"Summary data saved as {summary_file}")

    # Save the data to a single Excel file with multiple sheets
    output_file_path = os.path.join(OUTPUT_DIR, 'Hourly_Billing_Activity_Summary_Formatted_Updated.xlsx')
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        # Save initial data with categories and automation ease
        data.to_excel(writer, sheet_name='Initial Data', index=False)

        # Save grouped data
        grouped.to_excel(writer, sheet_name='Grouped Data', index=False)

        # Save summary data
        summary.to_excel(writer, sheet_name='Summary Data', index=False)

        # Adding filters and formatting
        for sheet_name in ['Initial Data', 'Grouped Data', 'Summary Data']:
            ws = writer.sheets[sheet_name]
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Add filters to all columns
            ws.auto_filter.ref = ws.dimensions

            # Center align all columns
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Format the headers
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font

        # Add a note explaining the automation ease ranking
        ws = writer.sheets['Initial Data']
        ws['A1'] = "Automation Ease Ranking: 5 = Easily Automated, 1 = Not Easily Automated"
        ws['A1'].font = Font(bold=True)

    logging.info(f"Output file saved as {output_file_path}")

    # Debug: Check for any rows in the original data that are not included in the grouping
    original_hours = data['Hours(Duration)'].sum()
    grouped_hours = grouped['Hours(Duration)'].sum()
    if original_hours != grouped_hours:
        missing_hours = original_hours - grouped_hours
        logging.warning(f"Missing hours: {missing_hours}")
        missing_data = data[~data.set_index(['Client Name', 'Category']).index.isin(grouped.set_index(['Client Name', 'Category']).index)]
        logging.warning("Rows in the original data that are not included in the grouping:")
        logging.warning(missing_data)
        missing_data.to_excel(os.path.join(OUTPUT_DIR, 'Missing_Data_Updated.xlsx'), index=False)
    else:
        logging.info("All hours are accounted for in the grouping.")

    # Dashboard
    # Creating a dashboard using Power BI or Tableau
    # (This step requires the use of external tools like Power BI or Tableau. Export the data as CSV to use in these tools.)

    # Save data as CSV for use in Power BI or Tableau
    data.to_csv(os.path.join(OUTPUT_DIR, 'Initial_Data_With_Categories_And_Automation_Ease.csv'), index=False)
    grouped.to_csv(os.path.join(OUTPUT_DIR, 'Grouped_Data.csv'), index=False)
    summary.to_csv(os.path.join(OUTPUT_DIR, 'Summary_Data_Updated.csv'), index=False)

    logging.info("Data saved as CSV for use in Power BI or Tableau.")
    logging.info("Script execution completed successfully.")

if __name__ == "__main__":
    main()

